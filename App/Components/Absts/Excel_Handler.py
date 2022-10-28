import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font


class Excel_Handler:
   Excel_Handler_Inst_lst = []
   # Class constructor.
   # The scope of this class is to have a data buffer, designed ass generic as possible to support multiple data elements.
   # Each instance of the class reprezents a line from the excel sheet (imported data or data to be written)
   # The mapping of the elements to the excel sheets is done as follows: el1 is the first column, el2 is the second column and so on.
   # In oder to prevent data overlap during multiple read/ write opperations, the data should be deleted once used.
   def __init__(self, el1=None, el2=None, el3=None, el4=None, el5=None, el6=None, el7=None):
      self.el1 = el1
      self.el2 = el2
      self.el3 = el3
      self.el4 = el4
      self.el5 = el5
      self.el6 = el6
      self.el7 = el7
      if self.el1 == None and self.el2 == None and self.el3 == None and self.el4 == None and self.el5 == None and self.el6 == None and self.el7 == None:
         pass
      else:
         Excel_Handler.Excel_Handler_Inst_lst.append(self)


   # Adds an element to the class
   @ classmethod
   def Add_Data(cls, el1, el2=None, el3=None, el4=None, el5=None, el6=None, el7=None):
      cls(el1, el2, el3, el4, el5, el6, el7)


   # Removes a class intance based on the given "cls_intance_index" parameter
   @classmethod
   def Remove_Data(cls, cls_intance_index):
      cls.Excel_Handler_Inst_lst.pop(cls_intance_index)


   # Removes all instances of the class
   @classmethod
   def Remove_All_Data(cls):
      cls.Excel_Handler_Inst_lst.clear()


   @classmethod
   # Writes all class intences to the given excel file and sheet.
   # filter_el and filter_val will filter the class data before writing. Only maching data will be written to the excel.
   # filter_el -> number of class parameter: if 1 -> el1, if 2 -> el2 and so on.
   # filter_val -> value to be filterd by(if u pass directly an ID)
   def Write_Sheet(cls, path, sheet, filter_el=None, filter_val=None):
      """
            A method that loads received data into excel file.
      """
      # Verify if the file exists.
      if os.path.exists(path):
         try:
            wb = load_workbook(path)
            if sheet in wb.sheetnames:
               wb.active = wb[sheet]
            else:
               wb.create_sheet(sheet)
               wb.active = wb[sheet]
            if filter_el == None and filter_val == None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "ID"
               ws["B1"] = "Comp Name"
               ws["C1"] = "Test Result(Pass/FAIL/SKIPP)"
               ws["D1"] = "Comments"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)

               row = 1
               col = 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el1)
                  row += 1
               row = 1
               col += 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el2)
                  row += 1
               row = 1
               col += 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el3)
                  row += 1
               row = 1
               col += 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el4)
                  row += 1
               row = 1
               col += 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el5)
                  row += 1
               row = 1
               col += 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el6)
                  row += 1
               row = 1
               col += 1
               for obj in cls.Excel_Handler_Inst_lst:
                  ws.cell(row+1, col, obj.el7)
                  row += 1
               row = 1
               col += 1
               wb.save(path)
               wb.close()
            elif filter_el != None and filter_val == None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "ID"
               ws["B1"] = "Comp Name"
               ws["C1"] = "Test Result(Pass/FAIL/SKIPP)"
               ws["D1"] = "Comments"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               ws["A2"] = cls.Excel_Handler_Inst_lst[filter_el].el1
               ws["B2"] = cls.Excel_Handler_Inst_lst[filter_el].el2
               ws["C2"] = cls.Excel_Handler_Inst_lst[filter_el].el3
               ws["D2"] = cls.Excel_Handler_Inst_lst[filter_el].el4
               ws["E2"] = cls.Excel_Handler_Inst_lst[filter_el].el5
               ws["F2"] = cls.Excel_Handler_Inst_lst[filter_el].el6
               ws["G2"] = cls.Excel_Handler_Inst_lst[filter_el].el7
               # Save and close
               wb.save(path)
               wb.close()
            elif filter_el == None and filter_val != None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "ID"
               ws["B1"] = "Comp Name"
               ws["C1"] = "Test Result(Pass/FAIL/SKIPP)"
               ws["D1"] = "Comments"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               for i in range(len(cls.Excel_Handler_Inst_lst)):
                  if str(cls.Excel_Handler_Inst_lst[i].el1) == str(filter_val):
                     ws["A2"] = cls.Excel_Handler_Inst_lst[i].el1
                     ws["B2"] = cls.Excel_Handler_Inst_lst[i].el2
                     ws["C2"] = cls.Excel_Handler_Inst_lst[i].el3
                     ws["D2"] = cls.Excel_Handler_Inst_lst[i].el4
                     ws["E2"] = cls.Excel_Handler_Inst_lst[i].el5
                     ws["F2"] = cls.Excel_Handler_Inst_lst[i].el6
                     ws["G2"] = cls.Excel_Handler_Inst_lst[i].el7

               # Save and close
               wb.save(path)
               wb.close()
            elif filter_el != None and filter_val != None:
               ws = wb.active
               # Delete everything from active sheet before writing to it.
               while ws.max_row > 1:
                  ws.delete_rows(2)
               # Create the headers.
               ws["A1"] = "ID"
               ws["B1"] = "Comp Name"
               ws["C1"] = "Test Result(Pass/FAIL/SKIPP)"
               ws["D1"] = "Comments"
               ws["A1"].font = Font(bold=True)
               ws["B1"].font = Font(bold=True)
               ws["C1"].font = Font(bold=True)
               ws["D1"].font = Font(bold=True)
               row = 1
               col = 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el1)
               col += 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el2)
               col += 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el3)
               col += 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el4)
               col += 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el5)
               col += 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el6)
               col += 1
               ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[filter_el].el7)
               col += 1
               col = 1
               row += 1
               for i in range(len(cls.Excel_Handler_Inst_lst)):
                  if str(cls.Excel_Handler_Inst_lst[i].el1) == str(filter_val):
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el1)
                     col += 1
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el2)
                     col += 1
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el3)
                     col += 1
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el4)
                     col += 1
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el5)
                     col += 1
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el6)
                     col += 1
                     ws.cell(row + 1, col, cls.Excel_Handler_Inst_lst[i].el7)
                     col = 1
                     row = 1
               # Save and close
               wb.save(path)
               wb.close()
         except Exception as e:
               return e
      # If file doesn't exist, create one.
      else:
         wb = Workbook()
         wb.remove_sheet(wb.active)
         wb.active = wb.create_sheet(sheet)
         ws = wb.active
         try:
            # Create the headers.
            ws["A1"] = "ID"
            ws["B1"] = "Comp Name"
            ws["C1"] = "Test Result(Pass/FAIL/SKIPP)"
            ws["D1"] = "Comments"
            ws["A1"].font = Font(bold=True)
            ws["B1"].font = Font(bold=True)
            ws["C1"].font = Font(bold=True)
            ws["D1"].font = Font(bold=True)
            row = 1
            col = 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el1)
               row += 1
            row = 1
            col += 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el2)
               row += 1
            row = 1
            col += 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el3)
               row += 1
            row = 1
            col += 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el4)
               row += 1
            row = 1
            col += 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el5)
               row += 1
            row = 1
            col += 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el6)
               row += 1
            row = 1
            col += 1
            for obj in cls.Excel_Handler_Inst_lst:
               ws.cell(row + 1, col, obj.el7)
               row += 1
            row = 1
            col += 1
            # Save and close
            wb.save(path)
            wb.close()
         except Exception as e:
            return e


   @classmethod
   def Import_Sheet(cls, path, sheet):
      """
      Get everything from the excel sheet and add each line to the class.
      Each column element shall be mapped as class's elements as follows:
      - column A = el1, column B = el2, etc.
      - row 0 = instance 0, row 1 = instance 1, etc.
      """
      if os.path.exists(path):
         wb = load_workbook(path)
         worksheets = wb.worksheets
         if wb[sheet] in worksheets:
            ws = wb.active
            wb.active = wb[sheet]
            for col in ws.iter_rows(min_row=2, min_col=ws.min_column, max_col=ws.max_column):
               if ws.max_column == 7:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value, col[6].value)
               elif ws.max_column == 6:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value, col[5].value)
               elif ws.max_column == 5:
                  cls(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value)
         else:
            return "Worksheet not found."
      else:
         return "There's no file at the given path."


   @classmethod
   def Get_Excel_Data(cls):
      return Excel_Handler.Excel_Handler_Inst_lst